import fpdf
from django.shortcuts import render
from django.db.models import Sum
from django.http import JsonResponse
from django.utils.text import slugify
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from project_management.models import Cutting, Material, TN50, TN71, TPK90, s100_rosi, TP120, UniversalFerez, \
    VerticalFerez, Kharzani, Drilling, LathingCnc, Carousel, StaticBalancing, PreAssembling, Welding, Casting, \
    Assembling, Painting, PackingDelivery, Outsourcing, Casting_Model
from django.shortcuts import render
from django.db.models import Sum
from django.http import JsonResponse
from django.shortcuts import render, redirect, reverse
from django.views.decorators.csrf import csrf_exempt
import json
import jdatetime
from persiantools.jdatetime import JalaliDate
from datetime import datetime
from project_management.models import MTO_Content, MTO_Headers
import operator
from django.shortcuts import HttpResponse
import pyodbc
from project_management.pdf_creator import create_mto_pdf
from django.http import FileResponse
from django.contrib.staticfiles.views import serve
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from persiantools.jdatetime import JalaliDate


# Create your views here.
@csrf_exempt
def home(request):
    return render(request, 'project_management/home.html')


@csrf_exempt
def print_mto(request):
    agenda_number = request.POST.get('title', )
    mto = MTO_Content.objects.filter(mto_headers__agenda_number=agenda_number)

    try:
        agenda_name = (mto[0].mto_headers.agenda_number.customer_name.first_name + " " + mto[
            0].mto_headers.agenda_number.customer_name.last_name)
    except:
        agenda_name = ""
    try:
        form_code = mto[0].mto_headers.form_code
    except:
        form_code = ""
    try:
        sp_code = mto[0].mto_headers.sp_code
    except:
        sp_code = ""
    try:
        collection = mto[0].mto_headers.collection
    except:
        collection = ""
    try:
        date = str(mto[0].mto_headers.date.year) + "-" + str(mto[0].mto_headers.date.month) + "-" + str(
            mto[0].mto_headers.date.day)
    except:
        date = ""
    try:
        machine_name = mto[0].mto_headers.machine_name
    except:
        machine_name = ""
    try:
        agenda_number = mto[0].mto_headers.agenda_number.agenda_number
    except:
        agenda_number = ""
    try:
        doc_no = mto[0].mto_headers.doc_no
    except:
        doc_no = ""
    try:
        producer = mto[0].mto_headers.producer
    except:
        producer = ""
    try:
        verified_by = mto[0].mto_headers.verified_by
    except:
        verified_by = ""
    try:
        machine_qdy = mto[0].mto_headers.machine_qdy
    except:
        machine_qdy = ""
    try:
        agenda_description = mto[0].mto_headers.agenda_description
    except:
        agenda_description = ""
    try:
        serial_no = mto[0].mto_headers.serial_no
    except:
        serial_no = ""

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    wb = openpyxl.load_workbook('C:/Users/Royal-ali/PycharmProjects/erp/static/1.xlsx')
    sheet = wb.active
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(13, 43):
        sheet.row_dimensions[i].height = 47

    for i in range(1, 42):
        for j in range(1, 38):
            if ((i == 2 and j == 5) or (i == 2 and j == 6)) or ((i == 1 and j == 5) or (i == 1 and j == 6)) or (
                    (i == 3 and j == 5) or (i == 3 and j == 6)) or ((i == 4 and j == 5) or (i == 4 and j == 6)) or (
                    (i == 5 and j == 5) or (i == 5 and j == 6)) or ((i == 6 and j == 5) or (i == 6 and j == 6)) or (
                    (i == 7 and j == 5) or (i == 7 and j == 6)) or ((i == 8 and j == 5) or (i == 8 and j == 6)):
                continue
            if ((i == 4 and j == 5) or (i == 5 and j == 5)) or ((i == 4 and j == 6) or (i == 5 and j == 6)) or (
                    (i == 4 and j == 7) or (i == 5 and j == 7)) or ((i == 4 and j == 8) or (i == 5 and j == 8)) or (
                    (i == 4 and j == 9) or (i == 5 and j == 9)) or ((i == 4 and j == 10) or (i == 5 and j == 10)) or (
                    (i == 4 and j == 11) or (i == 5 and j == 11)):
                continue
            if ((i == 9 and j == 1) or (i == 10 and j == 1)) or ((i == 9 and j == 2) or (i == 10 and j == 2)) or (
                    (i == 9 and j == 3) or (i == 10 and j == 3)) or ((i == 9 and j == 4) or (i == 10 and j == 4)):
                continue
            if i == 5 or i == 1 or i == 3 or i == 7:
                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                )
            if i == 6 or i == 2 or i == 4 or i == 8:
                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000'),
                )
            if i != 6 and i != 5 and i != 1 and i != 2 and i != 3 and i != 4 and i != 7 and i != 8:
                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000')
                )
            sheet.cell(row=i, column=j).border = thin_border

    sheet['A2'].value = agenda_number
    sheet['C2'].value = form_code
    sheet['A4'].value = sp_code
    sheet['C4'].value = machine_name
    sheet['A6'].value = collection
    sheet['C6'].value = agenda_number
    sheet['A8'].value = date
    sheet['C8'].value = doc_no
    sheet['B10'].value = producer
    sheet['C10'].value = verified_by
    sheet['E10'].value = machine_qdy

    img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/logo.png')
    img.anchor = 'L1'
    # sheet.add_image(img)
    # sheet.merge_cells('J9:Q9')
    # top_left_cell = sheet['Q9']
    # top_left_cell.value = agenda_description

    sheet.merge_cells('J9:Q10')
    cell = sheet.cell(row=9, column=10)
    cell.value = agenda_description
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrinkToFit=True)

    sheet.merge_cells('T10:U10')
    cell = sheet.cell(row=10, column=20)
    cell.value = serial_no
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrinkToFit=True)

    counter = 13
    for i in range(len(mto)):
        part_name = mto[i].part_name
        sps_part = mto[i].sps_part
        ver = mto[i].ver
        material_type = mto[i].material_type
        description = mto[i].description
        material = mto[i].material
        quantity = mto[i].quantity
        raw_material_dimension = mto[i].raw_material_dimension
        mass = mto[i].mass
        remark = mto[i].remark

        sheet['B' + str(counter)].value = part_name
        sheet['B' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['C' + str(counter)].value = sps_part
        sheet['C' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['D' + str(counter)].value = ver
        sheet['D' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['E' + str(counter)].value = material_type
        sheet['E' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['F' + str(counter)].value = description
        sheet['F' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['G' + str(counter)].value = material
        sheet['G' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['H' + str(counter)].value = quantity
        sheet['H' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['I' + str(counter)].value = raw_material_dimension
        sheet['I' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['J' + str(counter)].value = mass
        sheet['J' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)
        sheet['K' + str(counter)].value = remark
        sheet['K' + str(counter)].alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True,
                                                        wrap_text=True)

        # sheet['L' + str(counter)].value = part_name
        # sheet['M' + str(counter)].value = part_name
        # sheet['N' + str(counter)].value = part_name
        # sheet['O' + str(counter)].value = part_name
        # sheet['P' + str(counter)].value = part_name
        # sheet['Q' + str(counter)].value = part_name
        # sheet['R' + str(counter)].value = part_name
        # sheet['S' + str(counter)].value = part_name
        # sheet['T' + str(counter)].value = part_name
        # sheet['U' + str(counter)].value = part_name
        counter += 1
    name = str(agenda_number) + ".xlsx"
    wb.save("C:/Users/Royal-ali/PycharmProjects/erp/static/" + name)
    # return HttpResponse
    # return FileResponse(pdf, as_attachment=True, filename='hello.pdf')
    return serve(request, name)


def get_buy_list(serial_no):
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=DESKTOP-HN31M9C\MSSQLSERVER4;'
                          'Database=SPS_Acc;'
                          'Trusted_Connection=yes;')

    class CursorByName():
        def __init__(self, cursor):
            self._cursor = cursor

        def __iter__(self):
            return self

        def __next__(self):
            row = self._cursor.__next__()

            return {description[0]: row[col] for col, description in enumerate(self._cursor.description)}

    serial = serial_no
    cursor = conn.cursor()
    cursor2 = conn.cursor()
    cursor.execute("SELECT * FROM [SPS_Acc].[dbo].[Supp_ReqHdr] WHERE Atf = '{}'".format(serial))
    headers = []
    for row in CursorByName(cursor):
        headers.append(row)

    reqitems_confirmed = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}
    reqitems_finished = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}
    reqitems_revoked = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}
    reqitems_Initial = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}

    for item in headers:
        cursor.execute("SELECT * FROM [SPS_Acc].[dbo].[Supp_ReqItem] WHERE HeaderRef = '{}'".format(item['Hdr_IDGen']))
        if item['Status'] == 3:
            for row in CursorByName(cursor):
                reqitems_finished[str(item['Inv_Code'])].append(row)
        if item['Status'] == 1:
            for row in CursorByName(cursor):
                reqitems_confirmed[str(item['Inv_Code'])].append(row)
        if item['Status'] == 0:
            for row in CursorByName(cursor):
                reqitems_Initial[str(item['Inv_Code'])].append(row)
        if item['Status'] == 2:
            for row in CursorByName(cursor):
                reqitems_revoked[str(item['Inv_Code'])].append(row)

    parts_confirmed = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}
    parts_finished = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}
    parts_revoked = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}
    parts_Initial = {'0': [], '6': [], '1': [], '10': [], '4': [], '16': [], '8': []}

    for key, val in reqitems_Initial.items():
        for item in val:
            cursor.execute("SELECT * FROM [SPS_Acc].[dbo].[Part] WHERE Serial = '{}'".format(item['PartRef']))
            for row in CursorByName(cursor):
                parts_Initial[key].append(row)

    for key, val in reqitems_confirmed.items():
        for item in val:
            cursor.execute("SELECT * FROM [SPS_Acc].[dbo].[Part] WHERE Serial = '{}'".format(item['PartRef']))
            for row in CursorByName(cursor):
                parts_confirmed[key].append(row)

    for key, val in reqitems_finished.items():
        for item in val:
            cursor.execute("SELECT * FROM [SPS_Acc].[dbo].[Part] WHERE Serial = '{}'".format(item['PartRef']))
            for row in CursorByName(cursor):
                parts_finished[key].append(row)

    for key, val in reqitems_revoked.items():
        for item in val:
            cursor.execute("SELECT * FROM [SPS_Acc].[dbo].[Part] WHERE Serial = '{}'".format(item['PartRef']))
            for row in CursorByName(cursor):
                parts_revoked[key].append(row)

    return parts_confirmed, parts_Initial, parts_finished, parts_revoked


@csrf_exempt
def print_report(request):
    # print("this is print report called")
    agenda_number = request.POST.get('title', )
    # print(agenda_number)
    mto = MTO_Content.objects.filter(mto_headers__agenda_number=agenda_number)

    try:
        agenda_name = (mto[0].mto_headers.agenda_number.customer_name.first_name + " " + mto[
            0].mto_headers.agenda_number.customer_name.last_name)
    except:
        agenda_name = ""
    try:
        form_code = mto[0].mto_headers.form_code
    except:
        form_code = ""
    try:
        sp_code = mto[0].mto_headers.sp_code
    except:
        sp_code = ""
    try:
        collection = mto[0].mto_headers.collection
    except:
        collection = ""
    try:
        date = str(mto[0].mto_headers.date.year) + "-" + str(mto[0].mto_headers.date.month) + "-" + str(
            mto[0].mto_headers.date.day)
    except:
        date = ""
    try:
        machine_name = mto[0].mto_headers.machine_name
    except:
        machine_name = ""
    try:
        agenda_number = mto[0].mto_headers.agenda_number.agenda_number
    except:
        agenda_number = ""
    try:
        doc_no = mto[0].mto_headers.doc_no
    except:
        doc_no = ""
    try:
        producer = mto[0].mto_headers.producer
    except:
        producer = ""
    try:
        verified_by = mto[0].mto_headers.verified_by
    except:
        verified_by = ""
    try:
        machine_qdy = mto[0].mto_headers.machine_qdy
    except:
        machine_qdy = ""
    try:
        agenda_description = mto[0].mto_headers.agenda_description
    except:
        agenda_description = ""
    try:
        serial_no = mto[0].mto_headers.serial_no
    except:
        serial_no = ""
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    wb = openpyxl.load_workbook('C:/Users/Royal-ali/PycharmProjects/erp/static/report.xlsx')
    sheet = wb.active
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(11, 43):
        sheet.row_dimensions[i].height = 47
        sheet.merge_cells('A{}:B{}'.format(i, i))
        sheet.merge_cells('C{}:D{}'.format(i, i))

    for i in range(1, 42):
        for j in range(1, 31):
            if ((i == 2 and j == 5) or (i == 2 and j == 6)) or ((i == 1 and j == 5) or (i == 1 and j == 6)) or (
                    (i == 3 and j == 5) or (i == 3 and j == 6)) or ((i == 4 and j == 5) or (i == 4 and j == 6)) or (
                    (i == 5 and j == 5) or (i == 5 and j == 6)) or ((i == 6 and j == 5) or (i == 6 and j == 6)) or (
                    (i == 7 and j == 5) or (i == 7 and j == 6)) or ((i == 8 and j == 5) or (i == 8 and j == 6)):
                continue
            if ((i == 4 and j == 5) or (i == 5 and j == 5)) or ((i == 4 and j == 6) or (i == 5 and j == 6)) or (
                    (i == 4 and j == 7) or (i == 5 and j == 7)) or ((i == 4 and j == 8) or (i == 5 and j == 8)) or (
                    (i == 4 and j == 9) or (i == 5 and j == 9)) or ((i == 4 and j == 10) or (i == 5 and j == 10)) or (
                    (i == 4 and j == 11) or (i == 5 and j == 11)):
                continue
            if ((i == 9 and j == 1) or (i == 10 and j == 1)) or ((i == 9 and j == 2) or (i == 10 and j == 2)) or (
                    (i == 9 and j == 3) or (i == 10 and j == 3)) or ((i == 9 and j == 4) or (i == 10 and j == 4)):
                continue
            if i == 5 or i == 1 or i == 3 or i == 7:
                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                )
            if i == 6 or i == 2 or i == 4 or i == 8:
                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000'),
                )
            if i != 6 and i != 5 and i != 1 and i != 2 and i != 3 and i != 4 and i != 7 and i != 8:
                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000')
                )
            sheet.cell(row=i, column=j).border = thin_border

    # sheet['A2'].value = agenda_number
    # sheet['C2'].value = form_code
    # sheet['A4'].value = sp_code
    # sheet['C4'].value = machine_name
    # sheet['A6'].value = collection
    # sheet['C6'].value = agenda_number
    # sheet['A8'].value = date
    # sheet['C8'].value = doc_no
    # sheet['B10'].value = producer
    # sheet['C10'].value = verified_by
    # sheet['E10'].value = machine_qdy

    img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/logo.png')
    img.anchor = 'L1'
    # sheet.add_image(img)
    # sheet.merge_cells('J9:Q9')
    # top_left_cell = sheet['Q9']
    # top_left_cell.value = agenda_description

    fontStyle = Font(name="Calibri")
    sheet.merge_cells('A2:B2')
    cell = sheet.cell(row=2, column=1)
    cell.font = fontStyle
    try:
        cell.value = request.user.groups.all()[0].name
    except:
        cell.value = "Admin"
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('C4:D4')
    cell = sheet.cell(row=4, column=3)
    cell.font = fontStyle
    cell.value = machine_name
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('C2:D2')
    cell = sheet.cell(row=2, column=3)
    cell.font = fontStyle
    cell.value = form_code
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('A4:B4')
    cell = sheet.cell(row=4, column=1)
    cell.font = fontStyle
    cell.value = sp_code
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('C6:D6')
    cell = sheet.cell(row=6, column=3)
    cell.font = fontStyle
    cell.value = agenda_number
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('A6:B6')
    cell = sheet.cell(row=6, column=1)
    cell.font = fontStyle
    cell.value = serial_no
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('C8:D8')
    cell = sheet.cell(row=8, column=3)
    cell.font = fontStyle
    cell.value = doc_no
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('AE2:AF2')
    cell = sheet.cell(row=2, column=31)
    cell.font = fontStyle
    try:
        cell.value = request.user.groups.all()[0].name
    except:
        cell.value = "Admin"
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('AG4:AH4')
    cell = sheet.cell(row=4, column=33)
    cell.font = fontStyle
    cell.value = machine_name
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('AG2:AH2')
    cell = sheet.cell(row=2, column=33)
    cell.font = fontStyle
    cell.value = form_code
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('AE4:AF4')
    cell = sheet.cell(row=4, column=31)
    cell.font = fontStyle
    cell.value = sp_code
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('AG6:AH6')
    cell = sheet.cell(row=6, column=33)
    cell.font = fontStyle
    cell.value = agenda_number
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('AE6:AF6')
    cell = sheet.cell(row=6, column=31)
    cell.font = fontStyle
    cell.value = serial_no
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    sheet.merge_cells('AG8:AH8')
    cell = sheet.cell(row=8, column=33)
    cell.font = fontStyle
    cell.value = doc_no
    cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

    counter = 11
    for i in range(len(mto)):
        part_name = mto[i].part_name
        sps_part = mto[i].sps_part
        ver = mto[i].ver
        material_type = mto[i].material_type
        description = mto[i].description
        material = mto[i].material
        quantity = mto[i].quantity
        raw_material_dimension = mto[i].raw_material_dimension
        mass = mto[i].mass
        remark = mto[i].remark

        sheet.merge_cells('A{}:B{}'.format(counter, counter))
        cell = sheet.cell(row=counter, column=1)
        fontStyle = Font(name="B Nazanin", size=8)
        cell.font = fontStyle
        cell.value = part_name
        cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

        sheet.merge_cells('C{}:D{}'.format(counter, counter))
        cell = sheet.cell(row=counter, column=3)
        fontStyle = Font(name="Calibri", size=8)
        cell.font = fontStyle
        cell.value = sps_part
        cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

        cell = sheet.cell(row=counter, column=5)
        fontStyle = Font(name="B Nazanin", size=7)
        cell.font = fontStyle
        object = mto[i].pp
        if object is not None:
            object = mto[i].pp.cutting.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != [] and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        # print(string)
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

            cell = sheet.cell(row=counter, column=6)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.tn50.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:
                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

            column = 7

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.tn71.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.tpk90.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.s100_rosi.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.tp120.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.universal_ferez.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.vertical_ferez.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.kharzani.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.drilling.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.lath_cnc.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.carousel.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.static_balancing.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.pre_assembling.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.welding.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:
                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.assembly.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.painting.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.packing_delivery.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.outsourcing.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        else:
                            cell.value = "در انتظار تایید کنترل کیفیت"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if not object[j].qc_approv:
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        else:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1
            #
            #
            #
            #
            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.casting_model.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].final_approv:
                            cell.value = "انجام شده"
                        else:
                            cell.value = "در انتظار تایید کنترل نهایی"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].final_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if not object[j].final_approv:
                            string += "در انتظار تایید کنترل نهایی" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        else:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1
            #
            cell = sheet.cell(row=counter, column=column)
            fontStyle = Font(name="B Nazanin", size=7)
            cell.font = fontStyle
            object = mto[i].pp.casting.all()
            if len(object) == 0:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                cell.fill = my_fill
                # cell.value = "ندارد"
            else:

                if len(object) == 1:
                    if object[0] is None:
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='darkGray')
                        cell.fill = my_fill
                        # cell.value = "ندارد"
                    else:
                        if object[0].qc_approv:
                            cell.value = "انجام شده"
                        elif (object[0].finish_actual is not None) and (not object[0].qc_approv):
                            cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[0].finish_actual is None:
                            cell.value = "انجام نشده"
                else:
                    string = ''
                    approv_list = []
                    not_finished = []
                    for j in range(len(object)):
                        if object[j].qc_approv:
                            approv_list.append(True)
                        else:
                            approv_list.append(False)
                        if (object[j].finish_actual is not None) and (not object[j].qc_approv):
                            string += "در انتظار تایید کنترل کیفیت" + str(j + 1)
                            # cell.value = "در انتظار تایید کنترل کیفیت"
                        elif object[j].finish_actual is None:
                            string += "انجام نشده" + str(j + 1)

                    if not (False in approv_list) and approv_list != []:
                        cell.value = "انجام شده"
                    else:
                        cell.value = string
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            column += 1

        cell = sheet.cell(row=counter, column=26)
        fontStyle = Font(name="B Nazanin", size=7)
        cell.font = fontStyle
        if mto[i].available_sps:

            img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/true.jpg')
            c2e = cm_to_EMU

            # Calculated number of cells width or height from cm into EMUs
            cellh = lambda x: c2e((x * 49.77) / 99)
            cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

            # Want to place image in row 5 (6 in excel), column 2 (C in excel)
            # Also offset by half a column.
            column = 25
            coloffset = cellw(0.18)
            row = counter - 1
            rowoffset = cellh(0.6)
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(40), p2e(40))

            marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            sheet.add_image(img)
        else:

            img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/false.jpg')
            c2e = cm_to_EMU

            # Calculated number of cells width or height from cm into EMUs
            cellh = lambda x: c2e((x * 49.77) / 99)
            cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

            # Want to place image in row 5 (6 in excel), column 2 (C in excel)
            # Also offset by half a column.
            column = 25
            coloffset = cellw(0.18)
            row = counter - 1
            rowoffset = cellh(0.6)
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(40), p2e(40))

            marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            sheet.add_image(img)
        cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

        cell = sheet.cell(row=counter, column=26)
        fontStyle = Font(name="B Nazanin", size=7)
        cell.font = fontStyle
        if mto[i].builtin_buy:

            img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/true.jpg')
            c2e = cm_to_EMU

            # Calculated number of cells width or height from cm into EMUs
            cellh = lambda x: c2e((x * 49.77) / 99)
            cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

            # Want to place image in row 5 (6 in excel), column 2 (C in excel)
            # Also offset by half a column.
            column = 26
            coloffset = cellw(0.18)
            row = counter - 1
            rowoffset = cellh(0.6)
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(40), p2e(40))

            marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            sheet.add_image(img)
        else:

            img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/false.jpg')
            c2e = cm_to_EMU

            # Calculated number of cells width or height from cm into EMUs
            cellh = lambda x: c2e((x * 49.77) / 99)
            cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

            # Want to place image in row 5 (6 in excel), column 2 (C in excel)
            # Also offset by half a column.
            column = 26
            coloffset = cellw(0.18)
            row = counter - 1
            rowoffset = cellh(0.6)
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(40), p2e(40))

            marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            sheet.add_image(img)
        cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

        cell = sheet.cell(row=counter, column=27)
        fontStyle = Font(name="B Nazanin", size=7)
        cell.font = fontStyle
        if mto[i].standard_buy:

            img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/true.jpg')
            c2e = cm_to_EMU

            # Calculated number of cells width or height from cm into EMUs
            cellh = lambda x: c2e((x * 49.77) / 99)
            cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

            # Want to place image in row 5 (6 in excel), column 2 (C in excel)
            # Also offset by half a column.
            column = 27
            coloffset = cellw(0.18)
            row = counter - 1
            rowoffset = cellh(0.6)
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(40), p2e(40))

            marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            sheet.add_image(img)
        else:
            img = openpyxl.drawing.image.Image('C:/Users/Royal-ali/PycharmProjects/erp/static/images/false.jpg')
            c2e = cm_to_EMU
            # Calculated number of cells width or height from cm into EMUs
            cellh = lambda x: c2e((x * 49.77) / 99)
            cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

            # Want to place image in row 5 (6 in excel), column 2 (C in excel)
            # Also offset by half a column.
            column = 27
            coloffset = cellw(0.18)
            row = counter - 1
            rowoffset = cellh(0.6)
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(40), p2e(40))

            marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            sheet.add_image(img)
        cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)

        counter += 1

    parts_confirmed, parts_initial, parts_finished, parts_revoked = get_buy_list(serial_no)
    # print(parts_confirmed, parts_initial, parts_finished, parts_revoked)
    counter = 11
    inv_map = {'1': 'انبار مرکزی', '10': 'انبار اقلام بدون شارژ', '3': 'انبار محصول', '6': 'انبار مواد خام'}
    for key, val in parts_confirmed.items():
        for i in range(len(val)):
            sheet.merge_cells('AE{}:AP{}'.format(counter, counter))
            thin_border = Border(
                left=Side(border_style=BORDER_THIN, color='00000000'),
                right=Side(border_style=BORDER_THIN, color='00000000'),
                top=Side(border_style=BORDER_THIN, color='00000000'),
                bottom=Side(border_style=BORDER_THIN, color='00000000')
            )
            sheet.row_dimensions[counter].height = 47
            cell = sheet.cell(row=counter, column=31)
            for k in range(31, 43):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = val[i]['Name']

            sheet.merge_cells('AQ{}:AZ{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=43)
            for k in range(43, 53):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            try:
                cell.value = inv_map[key]
            except:
                cell.value = "انبار شماره {}".format(key)
            sheet.merge_cells('BA{}:BF{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=53)
            for k in range(53, 59):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = "تایید شده"
            counter += 1
    for key, val in parts_initial.items():
        for i in range(len(val)):
            sheet.merge_cells('AE{}:AP{}'.format(counter, counter))
            thin_border = Border(
                left=Side(border_style=BORDER_THIN, color='00000000'),
                right=Side(border_style=BORDER_THIN, color='00000000'),
                top=Side(border_style=BORDER_THIN, color='00000000'),
                bottom=Side(border_style=BORDER_THIN, color='00000000')
            )
            sheet.row_dimensions[counter].height = 47
            cell = sheet.cell(row=counter, column=31)
            for k in range(31, 43):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = val[i]['Name']

            sheet.merge_cells('AQ{}:AZ{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=43)
            for k in range(43, 53):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            try:
                cell.value = inv_map[key]
            except:
                cell.value = "انبار شماره {}".format(key)

            sheet.merge_cells('BA{}:BF{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=53)
            for k in range(53, 59):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = "ثبت اولیه"
            counter += 1
    for key, val in parts_finished.items():
        for i in range(len(val)):
            sheet.merge_cells('AE{}:AP{}'.format(counter, counter))
            thin_border = Border(
                left=Side(border_style=BORDER_THIN, color='00000000'),
                right=Side(border_style=BORDER_THIN, color='00000000'),
                top=Side(border_style=BORDER_THIN, color='00000000'),
                bottom=Side(border_style=BORDER_THIN, color='00000000')
            )
            sheet.row_dimensions[counter].height = 47
            cell = sheet.cell(row=counter, column=31)
            for k in range(31, 43):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = val[i]['Name']

            sheet.merge_cells('AQ{}:AZ{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=43)
            for k in range(43, 53):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            try:
                cell.value = inv_map[key]
            except:
                cell.value = "انبار شماره {}".format(key)

            sheet.merge_cells('BA{}:BF{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=53)
            for k in range(53, 59):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = "خاتمه یافته"
            counter += 1
    for key, val in parts_revoked.items():
        for i in range(len(val)):
            sheet.merge_cells('AE{}:AP{}'.format(counter, counter))
            thin_border = Border(
                left=Side(border_style=BORDER_THIN, color='00000000'),
                right=Side(border_style=BORDER_THIN, color='00000000'),
                top=Side(border_style=BORDER_THIN, color='00000000'),
                bottom=Side(border_style=BORDER_THIN, color='00000000')
            )
            sheet.row_dimensions[counter].height = 47
            cell = sheet.cell(row=counter, column=31)
            for k in range(31, 43):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = val[i]['Name']

            sheet.merge_cells('AQ{}:AZ{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=43)
            for k in range(43, 53):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            try:
                cell.value = inv_map[key]
            except:
                cell.value = "انبار شماره {}".format(key)

            sheet.merge_cells('BA{}:BF{}'.format(counter, counter))
            cell = sheet.cell(row=counter, column=53)
            for k in range(53, 59):
                sheet.cell(row=counter, column=k).border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            fontStyle = Font(name="B Nazanin", size=20)
            cell.font = fontStyle
            cell.value = "ابطال گردیده"

            # for jj in range(44,100):
            #     try:
            #         sheet.merge_cells('BA{}:BF{}'.format(counter, counter))
            #         cell = sheet.cell(row=counter, column=jj)
            #         # for k in range(32, 44):
            #         #     sheet.cell(row=counter, column=k).border = thin_border
            #         cell.alignment = Alignment(horizontal='center', vertical='center', shrinkToFit=True, wrap_text=True)
            #         fontStyle = Font(name="B Nazanin", size=20)
            #         cell.font = fontStyle
            #         cell.value = val[i]['Name']
            #         print(jj)
            #     except:
            #         pass

            counter += 1

    name = str(agenda_number) + "_report.xlsx"
    wb.save("C:/Users/Royal-ali/PycharmProjects/erp/static/" + name)
    # return HttpResponse
    # return FileResponse(pdf, as_attachment=True, filename='hello.pdf')
    return serve(request, name)


@csrf_exempt
def inventory(request):
    title = request.POST.get('title', )

    table_names = ["Inv_Exit",
                   "Inv_RelationVch",
                   "Inv_ReportControlParts",
                   "Inv_Vat",
                   "Inv_VchConfirm",
                   "Inv_VchHdr",
                   "Inv_VchHdrVchNum",
                   "Inv_VchItem",
                   "Inv_VchItemPrice",
                   "Inv_VchItemPrice2",
                   "Inv_VchLogHdr",
                   "Inv_VchLogItem",
                   "Inv_VchLogItemPrice",
                   "Inv_VchPriceElements",
                   "Inv_VchPricingHdr",
                   "Inv_VchReason",
                   "Inv_VchTrans",
                   "InvAccCode",
                   "InvDef",
                   "InvDepotParts",
                   "InvDepotPeriod",
                   "InvReqPart_Hdr",
                   "InvReqPart_HdrLog",
                   "InvReqPart_Item",
                   "InvReqPart_ItemLog",
                   "InvReqPart_Status",
                   "InvReView",
                   "InvReView_RemindQty",
                   "InvVchCtgry",
                   "InvVchType",
                   "InvVchType_RelatedCaption",
                   ]

    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=DESKTOP-HN31M9C\MSSQLSERVER4;'
                          'Database=SPS_Acc;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()
    val = title
    for i in range(len(table_names)):
        try:
            command = "SELECT * FROM SPS_Acc.dbo." + table_names[i] + " WHERE Name LIKE '%{}%'".format(val)
            cursor.execute(command)
            x = cursor.fetchall()

        except:
            pass

    products = x
    context = {'products': products}
    return render(request, 'project_management/product_list.html', context)


@csrf_exempt
def report(request):
    agenda_number = request.POST.get('fname', )
    queue = request.POST.get('lname', )
    # print(agenda_number, queue)
    if (agenda_number == '' and queue == '') or (agenda_number is None and queue is None):

        mto_content = MTO_Content.objects.all()
        # contact = Contact.objects.select_related().filter(client=book.id)
        # addresse = contact.client.addresse
        mto_header = MTO_Headers.objects.all()

    elif agenda_number is not None and queue == '':
        # print("here2")
        # print(agenda_number)
        mto_content = MTO_Content.objects.filter(mto_headers__agenda_number__agenda_number=agenda_number)
        # print(mto_content)
    elif agenda_number == '' and queue is not None:
        # print("here3")
        # mto_header = MTO_Headers.objects.all()
        mto_content = MTO_Content.objects.filter(mto_headers__queue=queue)
    elif agenda_number is not None and queue is not None:
        # print("here4")
        # mto_header = MTO_Headers.objects.filter(agenda_number__agenda_number=agenda_number)
        mto_content = MTO_Content.objects.filter(mto_headers__queue=queue,
                                                 mto_headers__agenda_number__agenda_number=agenda_number)

    mto_content = sorted(mto_content, key=operator.attrgetter('mto_headers.agenda_number.agenda_number'))
    mto_content_nums = []
    mto_queue_nums = []
    for item in mto_content:
        total_cost_planning = []
        total_cost_actual = []

        if item.cutting is not None:
            if item.cutting.finish_actual is not None:
                total_cost_planning.append(float(item.cutting.unit_cost) * (
                        item.cutting.finish_planning - item.cutting.start_planning).total_seconds())
                total_cost_actual.append(float(item.cutting.unit_cost) * (
                        item.cutting.finish_actual - item.cutting.start_actual).total_seconds())

        if item.tn50 is not None:
            if item.tn50.finish_actual is not None:
                total_cost_planning.append(
                    float(item.tn50.total_cost) * (
                            item.tn50.finish_planning - item.tn50.start_planning).total_seconds())
                total_cost_actual.append(
                    float(item.tn50.unit_cost) * (item.tn50.finish_actual - item.tn50.start_actual).total_seconds())

        if item.tn71 is not None:
            if item.tn71.finish_actual is not None:
                total_cost_planning.append(
                    float(item.tn71.total_cost) * (
                            item.tn71.finish_planning - item.tn71.start_planning).total_seconds())
                total_cost_actual.append(
                    float(item.tn71.unit_cost) * (item.tn71.finish_actual - item.tn71.start_actual).total_seconds())

        if item.tpk90 is not None:
            if item.tpk90.finish_actual is not None:
                total_cost_planning.append(
                    float(item.tpk90.total_cost) * (
                            item.tpk90.finish_planning - item.tpk90.start_planning).total_seconds())
                total_cost_actual.append(
                    float(item.tpk90.unit_cost) * (item.tpk90.finish_actual - item.tpk90.start_actual).total_seconds())

        if item.s100_rosi is not None:
            if item.s100_rosi.finish_actual is not None:
                total_cost_planning.append(float(item.s100_rosi.total_cost) * (
                        item.s100_rosi.finish_planning - item.s100_rosi.start_planning).total_seconds())
                total_cost_actual.append(float(item.s100_rosi.unit_cost) * (
                        item.s100_rosi.finish_actual - item.s100_rosi.start_actual).total_seconds())

        if item.tp120 is not None:
            if item.tp120.finish_actual is not None:
                total_cost_planning.append(
                    float(item.tp120.total_cost) * (
                            item.tp120.finish_planning - item.tp120.start_planning).total_seconds())
                total_cost_actual.append(
                    float(item.tp120.unit_cost) * (item.tp120.finish_actual - item.tp120.start_actual).total_seconds())

        if item.universal_ferez is not None:
            if item.universal_ferez.finish_actual is not None:
                total_cost_planning.append(float(item.universal_ferez.total_cost) * (
                        item.universal_ferez.finish_planning - item.universal_ferez.start_planning).total_seconds())
                total_cost_actual.append(float(item.universal_ferez.unit_cost) * (
                        item.universal_ferez.finish_actual - item.universal_ferez.start_actual).total_seconds())

        if item.vertical_ferez is not None:
            if item.vertical_ferez.finish_actual is not None:
                total_cost_planning.append(float(item.vertical_ferez.total_cost) * (
                        item.vertical_ferez.finish_planning - item.vertical_ferez.start_planning).total_seconds())
                total_cost_actual.append(float(item.vertical_ferez.unit_cost) * (
                        item.vertical_ferez.finish_actual - item.vertical_ferez.start_actual).total_seconds())

        if item.kharzani is not None:
            if item.kharzani.finish_actual is not None:
                total_cost_planning.append(float(item.kharzani.total_cost) * (
                        item.kharzani.finish_planning - item.kharzani.start_planning).total_seconds())
                total_cost_actual.append(float(item.kharzani.unit_cost) * (
                        item.kharzani.finish_actual - item.kharzani.start_actual).total_seconds())

        if item.drilling is not None:
            if item.drilling.finish_actual is not None:
                total_cost_planning.append(float(item.drilling.total_cost) * (
                        item.drilling.finish_planning - item.drilling.start_planning).total_seconds())
                total_cost_actual.append(float(item.drilling.unit_cost) * (
                        item.drilling.finish_actual - item.drilling.start_actual).total_seconds())

        if item.lath_cnc is not None:
            if item.lath_cnc.finish_actual is not None:
                total_cost_planning.append(float(item.lath_cnc.total_cost) * (
                        item.lath_cnc.finish_planning - item.lath_cnc.start_planning).total_seconds())
                total_cost_actual.append(float(item.lath_cnc.unit_cost) * (
                        item.lath_cnc.finish_actual - item.lath_cnc.start_actual).total_seconds())

        if item.carousel is not None:
            if item.carousel.finish_actual is not None:
                total_cost_planning.append(float(item.carousel.total_cost) * (
                        item.carousel.finish_planning - item.carousel.start_planning).total_seconds())
                total_cost_actual.append(float(item.carousel.unit_cost) * (
                        item.carousel.finish_actual - item.carousel.start_actual).total_seconds())

        if item.static_balancing is not None:
            if item.static_balancing.finish_actual is not None:
                total_cost_planning.append(float(item.static_balancing.total_cost) * (
                        item.static_balancing.finish_planning - item.static_balancing.start_planning).total_seconds())
                total_cost_actual.append(float(item.static_balancing.unit_cost) * (
                        item.static_balancing.finish_actual - item.static_balancing.start_actual).total_seconds())

        if item.pre_assembling is not None:
            if item.pre_assembling.finish_actual is not None:
                total_cost_planning.append(float(item.pre_assembling.total_cost) * (
                        item.pre_assembling.finish_planning - item.pre_assembling.start_planning).total_seconds())
                total_cost_actual.append(float(item.pre_assembling.unit_cost) * (
                        item.pre_assembling.finish_actual - item.pre_assembling.start_actual).total_seconds())

        if item.welding is not None:
            if item.welding.finish_actual is not None:
                total_cost_planning.append(float(item.welding.total_cost) * (
                        item.welding.finish_planning - item.welding.start_planning).total_seconds())
                total_cost_actual.append(float(item.welding.unit_cost) * (
                        item.welding.finish_actual - item.welding.start_actual).total_seconds())

        if item.assembly is not None:
            if item.assembly.finish_actual is not None:
                total_cost_planning.append(float(item.assembly.total_cost) * (
                        item.assembly.finish_planning - item.assembly.start_planning).total_seconds())
                total_cost_actual.append(float(item.assembly.unit_cost) * (
                        item.assembly.finish_actual - item.assembly.start_actual).total_seconds())

        if item.painting is not None:
            if item.painting.finish_actual is not None:
                total_cost_planning.append(float(item.painting.total_cost) * (
                        item.painting.finish_planning - item.painting.start_planning).total_seconds())
                total_cost_actual.append(float(item.painting.unit_cost) * (
                        item.painting.finish_actual - item.painting.start_actual).total_seconds())

        if item.packing_delivery is not None:
            if item.packing_delivery.finish_actual is not None:
                total_cost_planning.append(float(item.packing_delivery.total_cost) * (
                        item.packing_delivery.finish_planning - item.packing_delivery.start_planning).total_seconds())
                total_cost_actual.append(float(item.packing_delivery.unit_cost) * (
                        item.packing_delivery.finish_actual - item.packing_delivery.start_actual).total_seconds())

        machining_costs_planning = sum(total_cost_planning)
        machining_costs_actual = sum(total_cost_actual)
        item.total_cost_planning = str(float(item.material_cost) + machining_costs_planning)
        item.total_cost_actual = str(float(item.material_cost) + machining_costs_actual)
        mto_content_nums.append(item.mto_headers.agenda_number.agenda_number)
        mto_queue_nums.append(item.mto_headers.queue)

    context = {'mto_content': mto_content, 'mto_header': set(mto_content_nums), 'mto_queue': set(mto_queue_nums)}

    return render(request, 'project_management/report.html', context)
    # return render(request, 'project_management/report.html')


@csrf_exempt
def gantt_chart_data(request):
    machine_names = {"cutting": Cutting, "tn50": TN50, "tn71": TN71, "tpk90": TPK90, "s100_rosi": s100_rosi,
                     "tp120": TP120, "universal_ferez": UniversalFerez, "vertical_ferez": VerticalFerez,
                     "kharzani": Kharzani, "drilling": Drilling, "lath_cnc": LathingCnc, "carousel": Carousel,
                     "static_balancing": StaticBalancing, "pre_assembling": PreAssembling, "welding": Welding,
                     "casting": Casting, "assembly": Assembling, "painting": Painting,
                     "packing_delivery": PackingDelivery, "outsourcing": Outsourcing, "casting_model": Casting_Model}
    data = json.loads(request.body)
    machine_name = data["machinename"]
    try:
        date = data["date"]
        date = jdatetime.date(int(date.split("-")[0]), int(date.split("-")[1]), int(date.split("-")[2])).togregorian()
    except:
        date = datetime.today()
    print(machine_name, date)

    query = machine_names[machine_name].objects.filter(start_planning__year=str(date.year),
                                                       start_planning__month=str(date.month),
                                                       start_planning__day=str(date.day))
    labels = []
    data = []
    # print(len(query))
    print(query)
    for entry in query.iterator():
        part_code = entry.part_code
        part_name = MTO_Content.objects.filter(sps_part=part_code)[0].part_name
        mto_headers_id = MTO_Content.objects.filter(sps_part=part_code)[0].mto_headers_id
        agenda_number = MTO_Headers.objects.filter(id=mto_headers_id)[0].agenda_number.agenda_number
        labels.append((part_code, agenda_number))
        time = str(entry.finish_planning - entry.start_planning).split(":")
        # print(int(time[0]) * 60 + int(time[1]))
        data.append((entry.start_planning.hour, entry.start_planning.minute, entry.finish_planning.hour,
                     entry.finish_planning.minute))
    # print(labels)
    # print(data)
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })


@csrf_exempt
def gantt_chart(request):
    machine_name = request.POST.get('machinename', )
    context = {'machine_name': machine_name}
    return render(request, 'project_management/Gantt.html', context)


@csrf_exempt
def time_chart(request):
    labels = []
    data_actual = []
    data_planning = []
    gregorian_date = ""
    try:
        data = json.loads(request.body)
        from_year = data["from_year"]
        from_month = data["from_month"]
        from_day = data["from_day"]
        to_year = data["to_year"]
        to_month = data["to_month"]
        to_day = data["to_day"]
        mat = data["material"]
        if to_year == '':
            to_year = jdatetime.datetime.today().year
        if to_month == '':
            to_month = jdatetime.datetime.today().month
        if to_day == '':
            to_day = jdatetime.datetime.today().day + 1
        if from_year == '':
            from_year = jdatetime.datetime.today().year
        if from_month == '':
            from_month = jdatetime.datetime.today().month
        if from_day == '':
            from_day = jdatetime.datetime.today().day

        gregorian_date_start = jdatetime.date(int(from_year), int(from_month), int(from_day))
        gregorian_date_end = jdatetime.date(int(to_year), int(to_month), int(to_day))

        if mat != '':
            query = Cutting.objects.filter(start_actual__range=[gregorian_date_start, gregorian_date_end])
            query = query.filter(material__material_no=mat)


        else:
            query = Cutting.objects.filter(start_actual__range=[gregorian_date_start, gregorian_date_end])

    except:
        query = Cutting.objects.all()

    for entry in query.iterator():
        labels.append(str(jdatetime.date(day=entry.start_actual.day,
                                         month=entry.start_actual.month,
                                         year=entry.start_actual.year)) + "\n" +
                      str(entry.start_actual.hour) + ":" + str(entry.start_actual.minute) + ":" +
                      str(entry.start_actual.second))

        delta_actual = entry.finish_actual - entry.start_actual
        delta_actual = delta_actual.total_seconds()
        data_actual.append(delta_actual)

        delta_planning = entry.finish_planning - entry.start_planning
        delta_planning = delta_planning.total_seconds()
        data_planning.append(delta_planning)
    data = [data_actual, data_planning]
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })
